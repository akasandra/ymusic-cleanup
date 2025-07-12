
import logging
from yandex_music import Client
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

logging.basicConfig(
    level=logging.WARN,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

token = open('token.txt').read().rstrip('\n')
client = Client(token, language='en').init()

def is_english(text: str) -> bool:
    if not text:
        return False
    # This regex matches any non-English character (outside basic Latin ranges)
    # Adjust pattern to allow spaces and basic punctuation if needed
    return not bool(re.findall(r"[^\u0000-\u007F]+", text))

def is_russian_genre(text: str) -> bool:
    if not text:
        return False
    return 'rus' in text or 'phonk' in text or 'local' in text

def get_online_data():
    
    # Get list of all liked tracks
    tracks = client.users_likes_tracks()
    print('Number of liked tracks: %d' % len(tracks))

    # Get list of all liked albums
    albums = client.users_likes_albums()
    print('Number of liked albums: %d' % len(albums))

    # Get list of all liked artists
    artists = client.users_likes_artists()
    print('Number of liked artists: %d' % len(artists))

    # Sort tracks by most recently 'liked'
    tracks = sorted(tracks, key=lambda item: (item.timestamp, item.album_id), reverse=True)

    # Sort albums by most recently 'liked'
    albums = sorted(albums, key=lambda item: (item.timestamp, item.album.artists[0].id if item.album.artists else 0), reverse=True)

    # Sort artists by most recently 'liked'
    artists = sorted(artists, key=lambda item: item.timestamp, reverse=True)

    return {
        'artists': artists,
        'albums': albums,
        'tracks': tracks,
    }

# google sheets bug: turns int fields into floats, parsed as X.0 instead of X
def strip_trailing_dot_zero(value):
    if value == None:
        return None
    s = str(value)
    if s.endswith('.0'):
        return s[:-2]  # remove the '.0' suffix
    return s

def load_changes(filename='./changes.xlsx') -> dict:
    """
    Reads Excel file with checkbox (boolean) in column A and string in column B.
    Returns list of dicts: [{False: "string1"}, {True: "string2"}, ...]
    """
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    
    changes = []
    for row in ws.iter_rows(min_row=2, max_col=11):
        like_on = row[0].value
        artist_id = row[1].value
        album_id = strip_trailing_dot_zero(row[2].value)
        track_id = row[3].value
        timestamp = row[4].value
        artist = row[5].value
        genres = row[6].value
        album = row[7].value
        track = row[8].value
        year = row[9].value
        genre = row[10].value
        
        # Normalize boolean value: openpyxl may read Excel TRUE/FALSE as str or bool
        if isinstance(like_on, str):
            like_on = like_on.strip().upper() == 'TRUE'
        else:
            like_on = bool(like_on)
        
        # If string cell is None, convert to empty string for consistency
        artist_id = str(artist_id) if artist_id is not None else ""
        album_id = str(album_id) if album_id is not None else ""
        track_id = str(track_id) if track_id is not None else ""
        timestamp = str(timestamp) if timestamp is not None else ""
        artist = str(artist) if artist is not None else ""
        genres = str(genres) if genres is not None else ""
        album = str(album) if album is not None else ""
        track = str(track) if track is not None else ""
        year = str(year) if year is not None else ""
        genre = str(genre) if genre is not None else ""
        
        changes.append({
            'artist_id': artist_id,
            'album_id': album_id,
            'track_id': track_id,
            'timestamp': timestamp,
            'like_on': like_on,
            'artist': artist,
            'genres': genres,
            'album': album,
            'track': track,
            'year': year,
            'genre': genre,
        })
    
    return changes

def update_changes(online_data, changes) -> list:
    """
    Populate changes with missing information from online_data (new tracks, changed info)
    """

    # Add missing artists
    for artist in online_data.get('artists'):
        artist_id = artist.artist.id
        genres = ', '.join(artist.artist.genres) if artist.artist.genres else ''
        name = artist.artist.name if artist.artist.name else ''

        match = next(
            (d for d in changes if d['artist_id'] == artist_id and not d['album_id'] and not d['track_id']),
            None
        )
        if not match:
            changes.append({
                'artist_id': artist_id,
                'album_id': None,
                'track_id': None,
                'timestamp': artist.timestamp,
                'like_on': True,
                'artist': name,
                'genres': genres,
                'genre': artist.artist.genres[0] if artist.artist.genres else ''
            })

    # Add missing albums
    for album in online_data.get('albums'):
        album_id = album.album.id
        genre = album.album.genre
        title = album.album.title

        match = next(
            (d for d in changes if d['album_id'] == album_id and not d['track_id']),
            None
        )
        if not match:
            changes.append({
                'artist_id': album.album.artists[0].id if album.album.artists else '',
                'album_id': album_id,
                'track_id': None,
                'timestamp': album.timestamp,
                'like_on': True,
                'artist': album.album.artists[0].name if album.album.artists else '',
                'genre': genre if genre else '',
            })

    # Add missing tracks
    new_tracks = 0
    for track in online_data.get('tracks'):
        match = next(
            (d for d in changes if d['track_id'] == track.id),
            None
        )
        if not match:
            new_tracks +=1
            changes.append({
                'artist_id': None,
                'album_id': track.album_id,
                'track_id': track.id,
                'timestamp': track.timestamp,
                'like_on': True
            })

    print("New tracks: %d" % new_tracks)

    # Fetch track information
    track_ids = [i.get('track_id') for i in changes if i.get('track_id') and (
        not i.get('track') 
        or not i.get('artist')
    )]
    track_ids = list(dict.fromkeys(track_ids))
    tracks = []
    if track_ids:
        print('Need extra information for %d tracks' % len(track_ids))
        tracks = client.tracks(with_positions=False, track_ids=track_ids)

    # Fetch album information
    album_ids = [
        i.get('album_id')
        for i in changes
        if not i.get('track_id') and i.get('album_id')
        and not any(a.album.id == i.get('album_id') for a in online_data.get('albums'))
        and not any(track_in_album for track_in_album in tracks if any(a.id == i.get('album_id') for a in track_in_album.albums))
    ]
    album_ids = list(dict.fromkeys(album_ids))
    albums = []
    if album_ids:
        print('Need extra information for %d albums' % len(album_ids))
        albums = client.albums(album_ids=album_ids)

    # Substitute incomplete file data (changes) with new online data (albums, tracks)
    for idx, c in enumerate(changes):
        if c['track_id']:
            track = next((t for t in tracks if t.id == c['track_id']), None)
            if track:
                c['artist_id'] = str(track.artists[0].id)
                c['album_id'] = str(track.albums[0].id)
                c['track'] = track.title if not track.version else '%s (%s)' % (track.title, track.version)
        if c.get('album_id'):
            album = next((a for a in albums if a.id == c['album_id']), None)
            if not album:
                track = next((t for t in tracks if t.albums[0].id == c['album_id']), None)
                if track:
                    album = track.albums[0]
            if album:
                c['album'] = album.title if album.title else ''
                c['genre'] = album.genre if album.genre else ''
                c['year'] = next((y for y in ( album.original_release_year, album.year, album.release_date[:4] if album.release_date else '') if y), None)
        if c.get('artist_id'):
            artist = next((a.artists[0] for a in albums if a.artists and a.artists[0].id == c['artist_id']), None)
            if not artist:
                artist = next((t.artists[0] for t in tracks if t.artists[0].id == c['artist_id']), None)
            if not artist:
                artist = next((a.artist for a in online_data['artists'] if a.artist.id == c['artist_id']), None)
            if artist:
                c['artist'] = artist.name if artist.name else ''
                c['genres'] = ', '.join(artist.genres)

        changes[idx] = c

    return sorted(
        changes,
        key=lambda x: (
            0 if not x.get('track_id', '') else 1,
            0 if is_russian_genre(x.get('genres', '')) else 1,
            0 if is_russian_genre(x.get('genre', '')) else 1,
            1 if is_english(x.get('artist', '')) else 0,
            x.get('artist', '').lower(),
            0 if not x.get('album_id', '') else 1,
            x.get('genres', '').lower(),
            x.get('genre', '').lower()
        )
    )

def dump_changes(changes, filename='./changes.xlsx'):
    """
    Writes the changes list (list of dicts) back to Excel file.
    Adds one extra row at the top with hardcoded checkbox and string.
    Each row: checkbox (boolean) in column A, string in column B.
    """
    wb = Workbook()
    ws = wb.active

    # Header row
    ws.cell(row=1, column=1, value='like')
    ws.cell(row=1, column=2, value='artist_id')
    ws.cell(row=1, column=3, value='album_id')
    ws.cell(row=1, column=4, value='track_id')
    ws.cell(row=1, column=5, value='like_time')
    ws.cell(row=1, column=6, value='artist')
    ws.cell(row=1, column=7, value='genres')
    ws.cell(row=1, column=8, value='album')
    ws.cell(row=1, column=9, value='track')
    ws.cell(row=1, column=10, value='year')
    ws.cell(row=1, column=11, value='genre')

    # Write the changes
    for row_idx, change in enumerate(changes, start=2):
        ws.cell(row=row_idx, column=1, value=change.get('like_on'))
        ws.cell(row=row_idx, column=2, value=change.get('artist_id'))
        ws.cell(row=row_idx, column=3, value=change.get('album_id'))
        ws.cell(row=row_idx, column=4, value=change.get('track_id'))
        ws.cell(row=row_idx, column=5, value=change.get('timestamp'))
        ws.cell(row=row_idx, column=6, value=change.get('artist'))
        ws.cell(row=row_idx, column=7, value=change.get('genres'))
        ws.cell(row=row_idx, column=8, value=change.get('album'))
        ws.cell(row=row_idx, column=9, value=change.get('track'))
        ws.cell(row=row_idx, column=10, value=change.get('year'))
        ws.cell(row=row_idx, column=11, value=change.get('genre'))

    wb.save(filename)

def set_likes_changes(online_data, changes, filename='./changes.xlsx'):

    add_artists = []
    add_albums = []
    add_tracks = []

    rm_artists = []
    rm_albums = []
    rm_tracks = []

    off_changes = [c for c in changes if not c['like_on']]
    on_changes = [c for c in changes if c['like_on']]

    print('Like on ', len(on_changes))
    print('Like off', len(off_changes))

    for c in off_changes:
        if c.get('track_id'):
            track = next((t for t in online_data['tracks'] if str(t.id) == c['track_id']), None)
            if track:
                rm_tracks.append(c['track_id'])
        elif c.get('album_id'):
            album = next((a for a in online_data['albums'] if str(a.album.id) == str(c['album_id'])), None)
            if album:
                rm_albums.append(c['album_id'])
        elif c.get('artist_id'):
            artist = next((a for a in online_data['artists'] if str(a.artist.id) == c['artist_id']), None)
            if artist:
                rm_artists.append(c['artist_id'])

    for c in on_changes:
        if c.get('track_id'):
            track = next((t for t in online_data['tracks'] if str(t.id) == c['track_id']), None)
            if not track:
                add_tracks.append(c['track_id'])
        elif c.get('album_id'):
            album = next((a for a in online_data['albums'] if str(a.album.id) == str(c['album_id'])), None)
            if not album:
                add_albums.append(c['album_id'])
        elif c.get('artist_id'):
            artist = next((a for a in online_data['artists'] if str(a.artist.id) == c['artist_id']), None)
            if not artist:
                add_artists.append(c['artist_id'])

    print('Summary of likes to change online:')
    print('Remove likes: artists %d albums %d tracks %d' % (len(rm_artists), len(rm_albums), len(rm_tracks)))
    print('Add likes:    artists %d albums %d tracks %d' % (len(add_artists), len(add_albums), len(add_tracks)))

    if rm_tracks:
        client.users_likes_tracks_remove(track_ids=rm_tracks)
    if rm_albums:
        client.users_likes_albums_remove(album_ids=rm_albums)
    if rm_artists:
        client.users_likes_artists_remove(artist_ids=rm_artists)

    if add_tracks:
        client.users_likes_tracks_add(track_ids=add_tracks)
    if add_albums:
        client.users_likes_albums_add(album_ids=add_albums)
    if add_artists:
        client.users_likes_artists_add(artist_ids=add_artists)

    print('Indicates no error. Don\'t forget to update online_data to sync the changes')

