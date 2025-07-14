import os
from openpyxl import load_workbook, Workbook
from utility import iso_to_utc_timestamp, strip_trailing_dot_zero, value_to_bool
from source import Source
from table_helper import TableHelper

# ContextManager
class WorkbookContext:
    
    def __init__(self, wb):
        self._wb = wb

    def __enter__(self) -> Workbook:
        return self._wb

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._wb.close()

    def __getattr__(self, name):
        # Delegate attribute access to the wrapped workbook
        return getattr(self._wb, name)

class XlsxSource(Source, TableHelper):

    def __init__(self, filename: str):
        self.filename = filename

    def _open_truncate(self):
        return WorkbookContext(Workbook())

    def _open_update(self):
        if os.path.isfile(self.filename):
            wb = load_workbook(self.filename, data_only=False)
        else:
            wb = Workbook()
        return WorkbookContext(wb)

    def _bulk_read(self, wb, min_row: int, max_row: int, column_count: int) -> list:
        """
        Reads Excel file with changes library.
        Each row describes an artist, album or track. Like is a checkbox.
        """
        # Per each row, define how each cell value is post-processed (func) using a key in 'processors' 
        processors = self.get_read_processors()
        
        ws = wb.active

        # Read every row in range and return rows as key-value dicts
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=column_count):

            # Empty key-value for each column
            c = {k: '' for k in self.COLUMN_KEYS[:column_count]}

            # Fill with cell data and post-processed values
            for idx, cell in enumerate(row):
                key = self.COLUMN_KEYS[idx]
                c[key] = processors[key](cell.value)

            # Add unix time 'time' key to the each element
            timestamp = c.get('timestamp')
            c['time'] = iso_to_utc_timestamp(timestamp) if timestamp else 0
            
            yield c

    def _bulk_write(self, wb, min_row: int, changes: list, columns: list):
        """
        Writes the changes list (list of dicts) back to Excel file
        with updates/additions
        """

        # Processors per each column we may need or default
        processors = self.get_write_processors(columns)
        
        ws = wb.active

        # Write the changes
        for i, c in enumerate(changes):
            for column, key in enumerate(columns):
                if not key in c:
                    continue
                value = processors[key](c[key])
                ws.cell(row=min_row+i, column=column+1, value=value)

        wb.save(self.filename)

# End