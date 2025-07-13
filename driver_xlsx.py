from openpyxl import load_workbook, Workbook
from utility import iso_to_utc_timestamp, strip_trailing_dot_zero, value_to_bool

class XlsxFileDriver:

    # Order of table columns and key names mapping for rows
    COLUMN_KEYS = [
        'like_on',
        'artist_id',
        'album_id',
        'track_id',
        'timestamp',
        'artist',
        'genres',
        'album',
        'track',
        'year',
        'genre'
    ]

    # Transformations after read value for every cell value, by key
    READ_PROCESSORS = {
        'like_on': value_to_bool,
        'artist_id': strip_trailing_dot_zero,
        'album_id': strip_trailing_dot_zero,
        'track_id': strip_trailing_dot_zero,
        'year': strip_trailing_dot_zero
    }

    # Transformations before write value to openpyxl, by key
    WRITE_PROCESSORS = {}

    def __init__(self, filename: str):
        self.filename = filename

    def bulk_write(self, changes: list):
        wb = Workbook()
        ws = wb.active

        # Header row
        for i, key in enumerate(self.COLUMN_KEYS):
            ws.cell(row=1, column=i+1, value=key)

        # Write full table (assume metadata is present)
        self._bulk_write(ws=ws, min_row=2, changes=changes, columns=self.COLUMN_KEYS)

        wb.save(self.filename)
        wb.close()

    def bulk_read(self, no_metadata: bool=False) -> list:
        num_columns = 5 if no_metadata else len(self.COLUMN_KEYS)

        wb = load_workbook(self.filename, data_only=True)
        try:
            ws = wb.active
            read_items = self._bulk_read(ws=ws, min_row=2, max_row=None, column_count=num_columns)
            return list(read_items)
        finally:
            wb.close()

    def _bulk_read(self, ws, min_row: int, max_row: int, column_count: int) -> list:
        """
        Reads Excel file with changes library.
        Each row describes an artist, album or track. Like is a checkbox.
        """
        # Per each row, define how each cell value is post-processed (func) using a key in 'processors' 
        processors = self.READ_PROCESSORS.copy()
        for k in self.COLUMN_KEYS:
            if not k in processors:
                processors[k] = lambda value: str(value).strip() if value != None and str(value) else ''

        # Read every row in range and return rows as key-value dicts
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=column_count):

            # Empty key-value for each column
            c = {k: '' for k in self.COLUMN_KEYS[:column_count]}

            # Fill with cell data and post-processed values
            for idx, cell in enumerate(row):
                key = self.COLUMN_KEYS[idx]
                c[key] = processors[key](cell.value)

            # Add unix time 'time' key to the each element
            timestamp = c.get('timestamp')
            c['time'] = iso_to_utc_timestamp(timestamp) if timestamp else 0
            
            yield c

    def _bulk_write(self, ws, min_row: int, changes: list, columns: list):
        """
        Writes the changes list (list of dicts) back to Excel file
        with updates/additions
        """

        # Processors per each column we may need or default
        processors = {k: lambda v: v for k in columns}
        for k, f in self.WRITE_PROCESSORS:
            processors[k] = f

        # Write the changes
        for row, c in enumerate(changes):
            for column, key in enumerate(columns):
                if not key in c:
                    continue
                value = processors[key](c[key])
                ws.cell(row=min_row+row, column=column+1, value=value)

    def bulk_update(self, new_changes: list, old_changes: list=None):
        # Reopen workbook
        wb = load_workbook(self.filename, data_only=False)
        ws = wb.active

        # Read full current table with likes state to see if any needs update checkbox
        old_changes = self.bulk_read(no_metadata=True) if not old_changes else old_changes

        # Finds newest state per like id
        def get_new_state(c):
            for new in new_changes:
                if new['artist_id'] == c['artist_id'] and new['album_id'] == c['album_id'] and new['track_id'] == c['track_id']:
                    return new

        # Update likes on old changes (assume order is consistent)
        num_old_rows = 2 + len(old_changes)
        num_updated = 0
        for i, c in enumerate(old_changes):
            # For each of the existing table rows, get the updated state
            new = get_new_state(c)
            if not new:
                continue

            # For rows with updated like/timestamp, update the row
            if c['like_on'] != new['like_on'] or c['time'] != new['time']:
                num_updated += 1
                print('Update row', i+2)
                self._bulk_write(ws=ws, min_row=2+i, changes=[new], columns=['like_on', 'timestamp'])

        print('Rows updated:', num_updated)

        # The rest of changes are new, write to table
        # Write new table rows (assume metadata is present for new_changes)

        new_changes = new_changes[len(old_changes):]

        # Aggregate rows that are missing
        # Ensure not duplicating rows
        def find_old_entry(c):
            for old in old_changes:
                if all((old[k] == c[k] for k in ['artist_id', 'album_id', 'track_id'])):
                    yield old

        new_changes = list((c for c in new_changes if not any(find_old_entry(c))))

        self._bulk_write(ws=ws, min_row=num_old_rows, changes=new_changes, columns=self.COLUMN_KEYS)
        print('Rows added:', len(new_changes))

        wb.save(self.filename)
        wb.close()

# End