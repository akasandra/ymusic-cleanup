from openpyxl import load_workbook, Workbook
from datetime import datetime, timezone

def iso_to_utc_timestamp(iso_str: str) -> int:
    # Parse ISO 8601 string (Python 3.8 requires replacing 'Z' with '+00:00' if present)
    dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
    # Convert to UTC timezone if not already UTC
    dt_utc = dt.astimezone(timezone.utc)
    # Return Unix timestamp as int
    return int(dt_utc.timestamp())

# google sheets bug: turns int fields into floats, parsed as X.0 instead of X
def strip_trailing_dot_zero(value) -> str:
    if value == None:
        return None
    s = str(value)
    if s.endswith('.0'):
        return s[:-2]  # remove the '.0' suffix
    return s

class XlsxFileDriver:
    def __init__(self, filename: str):
        self.filename = filename

    def bulk_read(self) -> list:
        """
        Reads Excel file with changes library.
        Each row describes an artist, album or track. Like is a checkbox.
        """
        wb = load_workbook(self.filename, data_only=True)
        ws = wb.active

        current_time = int(datetime.now(timezone.utc).timestamp())
        
        changes = []
        for row in ws.iter_rows(min_row=2, max_col=11):
            like_on = row[0].value
            artist_id = strip_trailing_dot_zero(row[1].value)
            album_id = strip_trailing_dot_zero(row[2].value)
            track_id = strip_trailing_dot_zero(row[3].value)
            timestamp = row[4].value
            artist = row[5].value
            genres = row[6].value
            album = row[7].value
            track = row[8].value
            year = strip_trailing_dot_zero(row[9].value)
            genre = row[10].value
            
            # Normalize boolean value: openpyxl may read Excel TRUE/FALSE as str or bool
            if isinstance(like_on, str):
                like_on = like_on.strip().upper() == 'TRUE'
            else:
                like_on = bool(like_on)
            
            # If string cell is None, convert to empty string for consistency
            artist_id = str(artist_id) if artist_id is not None else ""
            album_id = str(album_id) if album_id is not None else ""
            track_id = str(track_id) if track_id is not None else ""
            timestamp = str(timestamp) if timestamp is not None else ""
            artist = str(artist) if artist is not None else ""
            genres = str(genres) if genres is not None else ""
            album = str(album) if album is not None else ""
            track = str(track) if track is not None else ""
            year = str(year) if year is not None else ""
            genre = str(genre) if genre is not None else ""
            
            changes.append({
                'artist_id': artist_id,
                'album_id': album_id,
                'track_id': track_id,
                'timestamp': timestamp,
                'like_on': like_on,
                'artist': artist,
                'genres': genres,
                'album': album,
                'track': track,
                'year': year,
                'genre': genre,
                'time': iso_to_utc_timestamp(timestamp) if timestamp else 0
            })
        
        return changes

    def bulk_write(self, changes: list):
        """
        Writes the changes list (list of dicts) back to Excel file
        with updates/additions
        """
        wb = Workbook()
        ws = wb.active

        # Header row
        ws.cell(row=1, column=1, value='like')
        ws.cell(row=1, column=2, value='artist_id')
        ws.cell(row=1, column=3, value='album_id')
        ws.cell(row=1, column=4, value='track_id')
        ws.cell(row=1, column=5, value='timestamp')
        ws.cell(row=1, column=6, value='artist')
        ws.cell(row=1, column=7, value='genres')
        ws.cell(row=1, column=8, value='album')
        ws.cell(row=1, column=9, value='track')
        ws.cell(row=1, column=10, value='year')
        ws.cell(row=1, column=11, value='genre')

        # Write the changes
        for row_idx, change in enumerate(changes, start=2):
            ws.cell(row=row_idx, column=1, value=change.get('like_on'))
            ws.cell(row=row_idx, column=2, value=change.get('artist_id'))
            ws.cell(row=row_idx, column=3, value=change.get('album_id'))
            ws.cell(row=row_idx, column=4, value=change.get('track_id'))
            ws.cell(row=row_idx, column=5, value=change.get('timestamp'))
            ws.cell(row=row_idx, column=6, value=change.get('artist'))
            ws.cell(row=row_idx, column=7, value=change.get('genres'))
            ws.cell(row=row_idx, column=8, value=change.get('album'))
            ws.cell(row=row_idx, column=9, value=change.get('track'))
            ws.cell(row=row_idx, column=10, value=change.get('year'))
            ws.cell(row=row_idx, column=11, value=change.get('genre'))

        wb.save(self.filename)

# End